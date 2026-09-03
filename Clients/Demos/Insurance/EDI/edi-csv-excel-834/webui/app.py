#!/usr/bin/env python3
"""EDI 834 CSV Excel Conversion — Demo Web UI."""

from __future__ import annotations

import csv
import io
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_file

app = Flask(__name__)

WEBUI_PORT = int(os.environ.get("WEBUI_PORT", "8140"))
ROUTES_DIR = Path(os.environ.get("ROUTES_DIR", "/routes"))
DOCUMENTS_DIR = Path(os.environ.get("DOCUMENTS_DIR", "/documents"))
SAMPLE_DIR = Path(os.environ.get("SAMPLE_DIR", "/samples"))
INBOUND_DIR = Path(os.environ.get("INBOUND_DIR", "/inbound"))
EDI_INBOUND_DIR = Path(os.environ.get("EDI_INBOUND_DIR", "/edi-inbound"))
EDI_OUT_DIR = Path(os.environ.get("EDI_OUT_DIR", "/output/834"))
CSV_OUT_DIR = Path(os.environ.get("CSV_OUT_DIR", "/output/csv"))
KICKOUT_DIR = Path(os.environ.get("KICKOUT_DIR", "/output/kickout"))
LAN_HINT = os.environ.get("LAN_HINT", "")
EIP_PUBLIC_URL = os.environ.get("EIP_PUBLIC_URL", "http://127.0.0.1:8139/eip/")
ROUTE_PDF_NAME = os.environ.get("ROUTE_PDF_NAME", "EDI834_CSV_Excel_V2_Route_Diagrams.pdf")
CAPABILITY_PDF_NAME = os.environ.get("CAPABILITY_PDF_NAME", "EDI834_CSV_Excel_Capability_Brief.pdf")

_MODULE_ID = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)
_ROUTE_SLUG = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
_SAFE_NAME = re.compile(r"^[A-Za-z0-9._-]+\.(csv|txt|xlsx|edi|834)$", re.I)
_XSLT_SUFFIXES = {".xsl", ".xslt"}


def route_slug(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "-", name.strip().lower()).strip("-")


def list_text_files(directory: Path, pattern: str):
    if not directory.exists():
        return []
    files = sorted(directory.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    out = []
    for path in files:
        if path.name.startswith("."):
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        out.append(
            {
                "name": path.name,
                "mtime": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "content": text,
                "size": path.stat().st_size,
            }
        )
    return out


def discover_v2_routes():
    if not ROUTES_DIR.is_dir():
        return []
    out = []
    for path in sorted(ROUTES_DIR.iterdir(), key=lambda p: p.name.lower()):
        if path.is_dir() and (path / "route.v2.xml").is_file():
            out.append({"id": route_slug(path.name), "dir": path.name, "name": path.name})
    return out


def resolve_route_dir(route_id: str) -> Path | None:
    if not route_id or not _ROUTE_SLUG.match(route_id):
        return None
    for meta in discover_v2_routes():
        if meta["id"] == route_id:
            candidate = (ROUTES_DIR / meta["dir"]).resolve()
            try:
                candidate.relative_to(ROUTES_DIR.resolve())
            except ValueError:
                return None
            return candidate if (candidate / "route.v2.xml").is_file() else None
    return None


def xlsx_to_csv_bytes(raw: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = wb.active
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if c is None else str(c) for c in row])
    return buf.getvalue()


@app.get("/")
def index():
    return render_template("index.html", title="EDI 834 CSV Excel Conversion", has_xslt=True)


@app.get("/api/health")
def api_health():
    return jsonify({"ok": True, "inbound": INBOUND_DIR.exists(), "ediOut": EDI_OUT_DIR.exists()})


@app.get("/api/samples")
def api_samples():
    files = []
    seen: set[str] = set()
    if SAMPLE_DIR.is_dir():
        for path in sorted(SAMPLE_DIR.rglob("*")):
            if not path.is_file():
                continue
            if path.suffix.lower() not in {".csv", ".txt", ".edi", ".834", ".xlsx"}:
                continue
            if path.name in seen:
                continue
            seen.add(path.name)
            item = {"name": path.name, "kind": path.suffix.lower().lstrip(".")}
            if path.suffix.lower() != ".xlsx":
                item["content"] = path.read_text(encoding="utf-8", errors="replace")
            files.append(item)
    return jsonify({"ok": True, "files": files})


@app.post("/api/inject")
def api_inject():
    payload = request.get_json(force=True, silent=True) or {}
    sample = (payload.get("sample") or "").strip()
    raw = payload.get("text") or payload.get("csv") or ""
    name = (payload.get("fileName") or "").strip()
    direction = (payload.get("direction") or "forward").strip().lower()

    if sample:
        if not _SAFE_NAME.match(sample):
            return jsonify({"ok": False, "error": "Invalid sample name"}), 400
        path = SAMPLE_DIR / sample
        if not path.is_file():
            hits = [p for p in SAMPLE_DIR.rglob(sample) if p.is_file() and p.name == sample]
            path = hits[0] if hits else path
        if not path.is_file():
            return jsonify({"ok": False, "error": "Unknown sample"}), 400
        name = sample
        if path.suffix.lower() == ".xlsx":
            raw = xlsx_to_csv_bytes(path.read_bytes())
            name = path.stem + ".csv"
        else:
            raw = path.read_text(encoding="utf-8", errors="replace")

    if not str(raw).strip():
        return jsonify({"ok": False, "error": "No file content"}), 400

    if not name or not _SAFE_NAME.match(name):
        stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        name = f"CUSTOM_{stamp}.csv"

    dest_dir = EDI_INBOUND_DIR if direction == "reverse" else INBOUND_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / name
    dest.write_text(raw if str(raw).endswith("\n") else str(raw) + "\n", encoding="utf-8")
    return jsonify({"ok": True, "file": dest.name, "dir": dest_dir.name})


@app.get("/api/results")
def api_results():
    return jsonify(
        {
            "ok": True,
            "edi": list_text_files(EDI_OUT_DIR, "*"),
            "csv": list_text_files(CSV_OUT_DIR, "*"),
            "kickout": list_text_files(KICKOUT_DIR, "*"),
        }
    )


@app.get("/api/wait-out")
def api_wait_out():
    kind = (request.args.get("kind") or "edi").strip()
    expected = (request.args.get("file") or "").strip()
    timeout = float(request.args.get("timeout", "60"))
    folder = {"edi": EDI_OUT_DIR, "csv": CSV_OUT_DIR, "kickout": KICKOUT_DIR}.get(kind, EDI_OUT_DIR)
    stem = Path(expected).stem if expected else ""
    deadline = time.time() + timeout
    while time.time() < deadline:
        files = list_text_files(folder, "*")
        match = next((f for f in files if Path(f["name"]).stem == stem), None) if stem else (files[0] if files else None)
        if match:
            return jsonify({"ok": True, "file": match, "files": files[:8]})
        time.sleep(1.0)
    return jsonify({"ok": False, "error": f"Timed out waiting for {kind} output"}), 504


@app.get("/api/v2/routes")
def api_v2_routes():
    return jsonify({"ok": True, "routes": discover_v2_routes()})


@app.get("/api/v2/routes/<route_id>/route.v2.xml")
def api_v2_route_xml(route_id: str):
    route_dir = resolve_route_dir(route_id)
    if not route_dir:
        return Response("Route not found", status=404, mimetype="text/plain")
    return Response((route_dir / "route.v2.xml").read_text(encoding="utf-8", errors="replace"), mimetype="application/xml")


@app.get("/api/v2/routes/<route_id>/diagram-groups.json")
def api_v2_diagram_groups(route_id: str):
    d = resolve_route_dir(route_id)
    if not d:
        return Response("Not found", status=404)
    path = d / "diagram-groups.json"
    if not path.is_file():
        return jsonify({"ok": True, "groups": []})
    return jsonify(json.loads(path.read_text(encoding="utf-8")))


@app.get("/api/v2/routes/<route_id>/modules/<module_id>.xml")
def api_v2_module_xml(route_id: str, module_id: str):
    if not _MODULE_ID.match(module_id):
        return Response("Invalid module id", status=400)
    route_dir = resolve_route_dir(route_id)
    if not route_dir:
        return Response("Route not found", status=404)
    path = route_dir / "modules" / f"{module_id}.xml"
    if not path.is_file():
        return Response("Module not found", status=404)
    return Response(path.read_text(encoding="utf-8", errors="replace"), mimetype="application/xml")


@app.get("/api/v2/xslt")
def api_xslt_list():
    if not ROUTES_DIR.is_dir():
        return jsonify({"ok": True, "files": []})
    out = []
    for path in sorted(ROUTES_DIR.rglob("*")):
        if path.is_file() and path.suffix.lower() in _XSLT_SUFFIXES:
            out.append({"path": path.relative_to(ROUTES_DIR).as_posix(), "name": path.name, "route": path.parent.name})
    return jsonify({"ok": True, "files": out})


@app.get("/api/v2/xslt/content")
def api_xslt_content():
    rel = (request.args.get("path") or "").strip()
    if not rel or Path(rel).is_absolute() or ".." in Path(rel).parts:
        return Response("not found", status=404)
    path = (ROUTES_DIR / rel).resolve()
    try:
        path.relative_to(ROUTES_DIR.resolve())
    except ValueError:
        return Response("not found", status=404)
    if not path.is_file():
        return Response("not found", status=404)
    return Response(path.read_text(encoding="utf-8", errors="replace"), mimetype="application/xml")


@app.context_processor
def _info_tab_standard_context():
    eip = (os.environ.get("EIP_PUBLIC_URL") or EIP_PUBLIC_URL).replace("localhost", "127.0.0.1")
    lan = os.environ.get("LAN_HINT") or LAN_HINT or ""
    urls = [{"label": "Local Web UI", "href": f"http://127.0.0.1:{WEBUI_PORT}/"}]
    if lan:
        urls.append({"label": "LAN Web UI", "href": lan})
    urls.append({"label": "EIP", "href": eip})
    return {
        "info_title": "EDI 834 CSV Excel Conversion",
        "info_blurb": "Drop CSV, tab-delimited TXT, or Excel (converted to CSV here). PilotFish normalizes to XML, maps enrollment rows to X12 834, and the reverse route writes CSV back out. Same beats as the public CSV / Excel / TXT to 834 tutorial — shorter, and no SNIP on this 23R1 stack.",
        "info_note": "Demo only — synthetic members, no SNIP. Excel Sheet Converter is 26R1; this 23R1 stack normalizes .xlsx in the Web UI.",
        "eip_url": eip,
        "lan_hint": lan,
        "info_urls": urls,
        "info_ports": [
            {"label": "Demo Web UI", "value": str(WEBUI_PORT)},
            {"label": "PilotFish EIP", "value": "8139"},
        ],
        "info_extra_links": [],
        "info_extra_sections": [],
        "test_results_pdf": None,
    }


try:
    from document_routes import (
        ensure_build_experience_api,
        ensure_build_replay_api,
        ensure_build_status_api,
        ensure_build_timing_api,
        ensure_document_routes,
    )
except ImportError:
    ensure_document_routes = None  # type: ignore
    ensure_build_timing_api = None  # type: ignore
    ensure_build_status_api = None  # type: ignore
    ensure_build_replay_api = None  # type: ignore
    ensure_build_experience_api = None  # type: ignore

if ensure_document_routes is not None:
    ensure_document_routes(
        app,
        DOCUMENTS_DIR,
        route_pdf_name=ROUTE_PDF_NAME,
        capability_pdf_name=CAPABILITY_PDF_NAME,
    )
if ensure_build_timing_api is not None:
    ensure_build_timing_api(app, DOCUMENTS_DIR)
if ensure_build_status_api is not None:
    ensure_build_status_api(app, DOCUMENTS_DIR)
if ensure_build_replay_api is not None:
    ensure_build_replay_api(app, DOCUMENTS_DIR)
if ensure_build_experience_api is not None:
    ensure_build_experience_api(app, DOCUMENTS_DIR)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=WEBUI_PORT, debug=False)
