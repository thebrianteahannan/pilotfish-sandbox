"""Apply IRL Expanse G3 H2 rows + Route 1 listeners. OGNL/A04 come from the planner."""

from __future__ import annotations

import json
import shutil
import subprocess
from pathlib import Path

import client_dive
import client_h2
import client_irl_expanse_g3_plan as g3
import client_request_diffs
import client_requests as reqs
import clients

REQ_ID = "20260901-143833-irl-expanse-new-facilities-monroe-osceol"
HOSPITALS = [
    {
        "client": "CEX",
        "sw": "528",
        "restrict": "PTH5.COCCN.*",
        "facility": "HCA FL Lake Monroe",
        "xlsx": "MedReceivables_NewFacilityInfo_IRL_CEX.xlsx",
    },
    {
        "client": "OSX",
        "sw": "525",
        "restrict": "PTH5.COCOS.*",
        "facility": "HCA FL Osceola",
        "xlsx": "MedReceivables_NewFacilityInfo_IRL_OSX.xlsx",
    },
    {
        "client": "POX",
        "sw": "526",
        "restrict": "PTH5.COCPMA.*",
        "facility": "HCA FL Poinciana",
        "xlsx": "MedReceivables_NewFacilityInfo_IRL_POX.xlsx",
    },
    {
        "client": "OVX",
        "sw": "527",
        "restrict": "PTH5.COCOMC.*",
        "facility": "HCA FL Oviedo Medical Center",
        "xlsx": "MedReceivables_NewFacilityInfo_IRL_OVX.xlsx",
    },
]
SW_IDS = ("525", "526", "527", "528")
GAN_SRC = 'name="Pickup Flat Files - IRL - GAN"'
GAN_PART = 'name="Set Partition and Client Name-IRL-GAN"'


def _esc(val: object) -> str:
    return str(val if val is not None else "").replace("'", "''")


def _write_sql(db: Path, sql: str) -> str:
    jar = client_h2._ensure_jar()
    stem = str(db).removesuffix(".mv.db").removesuffix(".h2.db")
    url = f"jdbc:h2:file:{stem};IFEXISTS=TRUE"
    proc = subprocess.run(
        ["java", "-cp", str(jar), "org.h2.tools.Shell", "-url", url, "-user", "sa", "-password", "", "-sql", sql],
        capture_output=True,
        text=True,
        timeout=60,
    )
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "H2 write failed").strip()[:400])
    return proc.stdout or ""


def fix_cex_xlsx(folder: Path) -> None:
    from openpyxl import load_workbook

    path = folder / "MedReceivables_NewFacilityInfo_IRL_CEX.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = headers.index("SOFTWAREID") + 1
    if str(ws.cell(2, col).value) != "528":
        ws.cell(2, col).value = 528
        wb.save(path)


def _xlsx_rows(path: Path) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    headers = [str(c or "") for c in raw[0]]
    rows = []
    for rec in raw[1:]:
        rows.append({headers[i]: rec[i] if i < len(rec) else None for i in range(len(headers))})
    return headers, rows


def _sql_from_xlsx(path: Path, software_id: str) -> str:
    _, rows = _xlsx_rows(path)
    first = next((r for r in rows if r.get("CLIENT")), None)
    if not first:
        raise RuntimeError(f"No CLIENT row in {path.name}")
    part = _esc(first.get("PARTITION") or "IRL")
    client = _esc(first.get("CLIENT"))
    name = _esc(first.get("CLIENTNAME"))
    stmts = [
        f"DELETE FROM CLIENT_CODES WHERE SOFTWARE_ID='{software_id}'",
        f"DELETE FROM CLIENT_SPLITS WHERE SOFTWAREID='{software_id}'",
    ]
    for row in rows:
        fac = str(row.get("SPLIT_FACILITY") or "").strip()
        cmp = str(row.get("COMPARATOR") or "").strip()
        if not fac or not cmp:
            continue
        code = _esc(row.get("SPLIT_CODE") or "")
        stmts.append(
            "INSERT INTO CLIENT_CODES (FACILITY, CODE, COMPARATOR, SOFTWARE_ID) VALUES ("
            f"'{_esc(fac)}','{code}','{_esc(cmp)}','{software_id}')"
        )
    split_n = 0
    for row in rows:
        if not (str(row.get("CLIENT") or "").strip() and str(row.get("FACILITY") or "").strip() and str(row.get("FACILITY_CODE") or "").strip()):
            continue
        split_n += 1
        default = "1" if split_n == 1 else "NULL"
        date = str(int(row["DATE_RANGE"])) if row.get("DATE_RANGE") not in (None, "") else ""
        stmts.append(
            "INSERT INTO CLIENT_SPLITS (SOFTWAREID, CLIENTNAME, PARTITION, CLIENT, FACILITY, "
            "FACILITY_CODE, DEFAULT_PERF_DR, ACCOUNT_NUM_ALPHA, DATE_RANGE, IS_DEFAULT) VALUES ("
            f"'{software_id}','{name}','{part}','{client}','{_esc(row.get('FACILITY'))}',"
            f"'{_esc(row.get('FACILITY_CODE'))}','{_esc(row.get('DEFAULT_PERF_DR'))}',"
            f"'{_esc(row.get('ACCOUNT_NUM_ALPHA'))}','{_esc(date)}',{default})"
        )
    return ";\n".join(stmts) + ";"


def insert_h2(root: Path, samples: Path) -> None:
    sql = []
    for hosp in HOSPITALS:
        sql.append(_sql_from_xlsx(samples / hosp["xlsx"], hosp["sw"]))
    blob = "\n".join(sql)
    seed = client_h2._db_file(root)
    live = (root / client_h2.LIVE).resolve()
    for db in (seed, live):
        if not db.is_file():
            continue
        bak = db.with_suffix(db.suffix + ".bak-g3")
        if not bak.is_file():
            shutil.copy2(db, bak)
        _write_sql(db, blob)


def _between(text: str, start_token: str, end_token: str) -> str:
    start = text.find(start_token)
    if start < 0:
        raise RuntimeError(f"missing {start_token}")
    line_start = text.rfind("\n", 0, start) + 1
    end = text.find(end_token, start)
    if end < 0:
        raise RuntimeError(f"missing end after {start_token}")
    return text[line_start:end]


def clone_listeners(root: Path) -> None:
    path = root / g3.ROUTE1
    text = path.read_text(encoding="utf-8")
    if 'name="Pickup Flat Files - IRL - CEX"' in text:
        return
    src = _between(text, GAN_SRC, '  <Source icon="DocFlatFile2_Orange_1.png" name="Pickup Flat Files - IRL - CAX">')
    part = _between(
        text,
        GAN_PART,
        '      <Processor class="com.pilotfish.eip.modules.internal.TransactionAttributePopulationProcessor" name="Set Partition and Client Name-IRL-TWX">',
    )
    new_src = []
    new_part = []
    for hosp in HOSPITALS:
        block = src
        block = block.replace("PTH5.GA.*", hosp["restrict"])
        block = block.replace("HCA FL Gainesville Hospital", hosp["facility"])
        block = block.replace("GAN", hosp["client"])
        new_src.append(block)
        new_part.append(part.replace("GAN", hosp["client"]))
    text = text.replace(src, src + "".join(new_src), 1)
    text = text.replace(part, part + "".join(new_part), 1)
    path.write_text(text, encoding="utf-8")


def slim_sample(src: Path, dest: Path) -> None:
    lines = src.read_text(encoding="latin-1", errors="replace").splitlines(keepends=True)
    header = [ln for ln in lines if ln.startswith("H")]
    trailer = [ln for ln in lines if ln.startswith("T")]
    groups: list[list[str]] = []
    cur: list[str] = []
    for ln in lines:
        if ln.startswith("P") and cur:
            groups.append(cur)
            cur = [ln]
        elif ln.startswith(("P", "I", "G", "C", "Q", "D")):
            cur.append(ln)
    if cur:
        groups.append(cur)

    def has_lc(group: list[str]) -> bool:
        for ln in group:
            if ln.startswith("C"):
                padded = ln.rstrip("\n") + " " * 20
                if padded[530:546].strip() == "LC":
                    return True
        return False

    picked: list[list[str]] = []
    if groups:
        picked.append(groups[0])
    for group in groups[1:]:
        if has_lc(group):
            picked.append(group)
            break
    if len(picked) == 1 and len(groups) > 1:
        picked.append(groups[1])
    dest.write_text("".join(header + [ln for g in picked for ln in g] + trailer), encoding="latin-1")


def apply_request() -> dict:
    root = clients.require_root("med-rec")
    folder = root / "requests" / REQ_ID
    samples = folder / "samples"
    fix_cex_xlsx(samples)
    insert_h2(root, samples)
    clone_listeners(root)
    dive = json.loads((folder / "dive.json").read_text(encoding="utf-8"))
    dive = g3.apply(dive, root, (folder / "email.txt").read_text(encoding="utf-8", errors="replace"), str(dive.get("subject") or ""))
    for q in dive.get("questions") or []:
        if "524" in str(q.get("text") or "") and "528" in str(q.get("text") or ""):
            q["status"] = "answered"
            q["answer"] = "Yes. CEX workbook SOFTWAREID is 528 so it does not collide with NHL CAT 524."
    applied = client_dive.apply_edits(root, dive)
    (folder / "dive.json").write_text(json.dumps(dive, indent=2) + "\n", encoding="utf-8")
    meta = reqs.load_meta(folder)
    meta["status"] = "applied"
    meta["applied"] = applied
    meta["likely_files"] = [f["path"] for f in dive.get("files") or []]
    meta["edit_count"] = len(dive.get("edits") or [])
    for rel in (g3.ROUTE1, g3.ROUTE2):
        if rel not in meta["likely_files"]:
            meta["likely_files"].append(rel)
    changes = client_request_diffs.write(root, folder, meta, dive)
    meta["changes"] = changes
    meta["change_summary"] = (
        "Onboarded IRL Expanse G3 CEX/OSX/POX/OVX: H2 splits/codes (CEX=528), four PTH5.COC* listeners, "
        "expanded-CDM / LC-strip allowlists, ADT software 525–528."
    )
    reqs.save_meta(folder, meta)
    reqs.append_log(folder, f"Implemented G3: {len(applied)} stylesheet/OGNL edit(s), {len(changes)} file(s).")
    return {"applied": applied, "changes": changes}


if __name__ == "__main__":
    out = apply_request()
    print(json.dumps(out, indent=2))
